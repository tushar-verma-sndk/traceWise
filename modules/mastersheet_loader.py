import openpyxl


def load_mastersheet(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    headers = {}
    platform_data = {}

    # Map header names
    for col in range(1, sheet.max_column + 1):
        headers[sheet.cell(row=1, column=col).value] = col

    required_columns = [
        "Platform Name",
        "Category",
        "Host Name",
        "Model",
        "Processor",
        "Chipset",
        "PCIe Speed",
        "Storage Controller",
        "Qty",
        "M.2 Slots",
        "Reference",
        "S0ix/S3"
    ]

    for row in range(2, sheet.max_row + 1):
        platform_name = sheet.cell(row=row, column=headers["Platform Name"]).value
        if not platform_name:
            continue

        platform_data[platform_name] = {
            "oem": sheet.cell(row=row, column=headers["Category"]).value,
            "host_name": sheet.cell(row=row, column=headers["Host Name"]).value,
            "model": sheet.cell(row=row, column=headers["Model"]).value,
            "processor": sheet.cell(row=row, column=headers["Processor"]).value,
            "chipset": sheet.cell(row=row, column=headers["Chipset"]).value,
            "pcie_speed": sheet.cell(row=row, column=headers["PCIe Speed"]).value,
            "storage_controller": sheet.cell(row=row, column=headers["Storage Controller"]).value,
            "qty": sheet.cell(row=row, column=headers["Qty"]).value,
            "m2_slots": sheet.cell(row=row, column=headers["M.2 Slots"]).value,
            "reference": sheet.cell(row=row, column=headers["Reference"]).value,
            "power_mode": sheet.cell(row=row, column=headers["S0ix/S3"]).value,
        }

    return platform_data